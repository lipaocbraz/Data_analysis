from openpyxl import load_workbook, Workbook

planilha = load_workbook('planilha_investimentos.xlsx')
pagina_renda_fixa = planilha['Mov Contas']

for linha in pagina_renda_fixa.iter_rows(values_only=True):
    print(linha[3])

#criando arquivo de texto
planilha_analise = Workbook()
pag1 = planilha_analise.active

with open ('dados_pos.txt', 'r', encoding='utf=8') as arquivo:
    for linha in arquivo:
        pag1.append(linha.split(', '))

planilha_analise.save('dados_pos.xlsx')

with open ('analise.txt', 'a', encoding='utf-8') as arquivo:
    for linha in pagina_renda_fixa.iter_rows(values_only=True):
        if linha[3] == 'Nubank':
            dado = str(linha)
            arquivo.write(f"{dado}\n")
    arquivo.close()

with open ('dados_pos.txt', 'a', encoding='utf-8') as arquivo:
    for linha in pagina_renda_fixa.iter_rows(values_only=True):
        try:
            verifica = float(linha[5])
        except TypeError:
            continue
        except ValueError:
            continue
        if verifica > 0:
            dado1 = str(linha)
            arquivo.write(f"{dado1}\n")